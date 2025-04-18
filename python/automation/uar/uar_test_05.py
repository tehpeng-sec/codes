import pandas as pd
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the Excel file
def filter_tesla_users(file_path, sheet_name='Sheet1'):
    # Load data
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # Add "Team" column and fill with "CRM APP"
    df['Team'] = "CRM APP"
    
    # Identify Tesla Users
    df['Tesla User'] = df['Target Account'].astype(str).str.contains("CA-KE-CRM-TSL-", na=False)
    
    # Save dataframe back to Excel
    output_file = "filtered_tesla_users.xlsx"
    df.to_excel(output_file, index=False, sheet_name=sheet_name)
    
    # Load the workbook to apply formatting
    wb = load_workbook(output_file)
    ws = wb[sheet_name]
    
    # Define highlight colors
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    
    # Get the column index of "Time"
    time_col_index = df.columns.get_loc("Time") + 1  # Adjust for Excel index (1-based)
    
    # Apply formatting for office hours and weekends
    for row_idx, time_value in enumerate(df['Time'], start=2):  # Excel rows start at 1, header is row 1
        if pd.notna(time_value):
            time_str = time_value.strftime('%H%M')  # Extract HHMM as string
            day_of_week = time_value.weekday()  # Monday=0, Sunday=6
            
            if int(time_str) < 830 or int(time_str) > 1800:
                ws.cell(row=row_idx, column=time_col_index).fill = red_fill  # Highlight red for outside office hours
            elif day_of_week in [5, 6]:  # Saturday or Sunday
                ws.cell(row=row_idx, column=time_col_index).fill = orange_fill  # Highlight orange for weekends
    
    # Save the updated workbook
    wb.save(output_file)
    print(f"Filtered data saved to {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Filter Tesla Users from Excel File")
    parser.add_argument("--source-file", required=True, help="Path to the source Excel file")
    args = parser.parse_args()
    
    filter_tesla_users(args.source_file)
