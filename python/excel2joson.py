import openpyxl
import json

def excel_to_json(excel_file, sheet_name):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb[sheet_name]

    # Create an empty list to store rows as dictionaries
    data = []

    # Iterate through rows in the sheet
    for row in sheet.iter_rows(values_only=True):
        # Assuming the first row contains column headers
        headers = row
        break

    # Iterate through rows (excluding the first row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Create a dictionary for each row
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        # Add row data to the list
        data.append(row_data)

    # Convert data list to JSON
    json_data = json.dumps(data, indent=4)

    return json_data

# Example usage
excel_file = 'data.xlsx'
sheet_name = 'Sheet1'
json_data = excel_to_json(excel_file, sheet_name)

# Write JSON data to a file
with open('data.json', 'w') as json_file:
    json_file.write(json_data)

print("Conversion successful. JSON data written to data.json.")
